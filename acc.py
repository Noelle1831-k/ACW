# ACC TPR FPR
import openpyxl
import os

if __name__ == "__main__":
	folder_paths = [
		"G/MBPP_3.5_G",
		"G/MBPP_4_G",
		"G/APPS_3.5_G",
		"G/APPS_4_G",
		"G/MBPP_Starcoder_G"
	]

	C = [
		"_1",
		"_2",
		"_3",
		"_4",
		"_5",
		"_6",
		"_over6"
	]
	ACCs = [0 for _ in range(30)]
	TPRs = [0 for _ in range(30)]
	FPRs = [0 for _ in range(30)]
	count = 0
	for folder_path in folder_paths:
		for c in C:
			FP = 0
			TP = 0
			TN = 0
			FN = 0
			new_folder_path = os.path.join(folder_path, folder_path[2:] + c)
			python_files = [f for f in os.listdir(new_folder_path) if f.endswith('.py')]
			for file in python_files:
				file_path = os.path.join(new_folder_path, file)
				try:
					with open(file_path, 'r') as file:
						content = file.read()
					compare_files = file_path[0] + '2' + file_path[1:]
					try:
						with open(compare_files, 'r') as compare_file:
							compare_content = compare_file.read()
						if content != compare_content:
							FN += 1
						else:
							TP += 1
					except Exception as e:
						print(f"Error processing file {compare_files}: {e}")
				except Exception as e:
					print(f"Error processing file {file_path}: {e}")

			new_folder_path = os.path.join('H', folder_path[2:-1] + 'H', folder_path[2:-1] + 'H' + c)
			python_files = [f for f in os.listdir(new_folder_path) if f.endswith('.py')]
			for file in python_files:
				file_path = os.path.join(new_folder_path, file)
				try:
					with open(file_path, 'r') as file:
						content = file.read()
					compare_files = file_path[0] + '2' + file_path[1:]
					try:
						with open(compare_files, 'r') as compare_file:
							compare_content = compare_file.read()
						if content != compare_content:
							TN += 1
						else:
							FP += 1
					except Exception as e:
						print(f"Error processing file {compare_files}: {e}")
				except Exception as e:
					print(f"Error processing file {file_path}: {e}")
			# print( new_folder_path,TP,TN,FP,FN)
			if TP + TN + FP + FN > 0:
				ACCs[count] = (TP + TN) / (TP + TN + FP + FN)
			else:
				ACCs[count] = 0
			if TP + FN > 0:
				TPRs[count] = TP / (TP + FN)
			else:
				TPRs[count] = 0
			if FP + TN > 0:
				FPRs[count] = FP / (FP + TN)
			else:
				FPRs[count] = 0
			count += 1
	workbook = openpyxl.Workbook()
	sheet = workbook.active
	sheet.title = 'detection'
	sheet['A1'] = 'Dataset'
	sheet['B1'] = 'ACC'
	sheet['C1'] = 'TPR'
	sheet['D1'] = 'FPR'
	for row_num, path in enumerate(folder_paths):
		sheet.cell(row=row_num + 2, column=1, value=path)
	for row_num, ACC in enumerate(ACCs):
		sheet.cell(row=row_num + 2, column=2, value=ACC)
	for row_num, TPR in enumerate(TPRs):
		sheet.cell(row=row_num + 2, column=3, value=TPR)
	for row_num, FPR in enumerate(FPRs):
		sheet.cell(row=row_num + 2, column=4, value=FPR)

	workbook.save('detection.xlsx')
